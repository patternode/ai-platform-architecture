"""
BNZ AI Use Case Prioritization Spreadsheet Generator
Creates a complete Excel workbook with all formulas and formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference
from datetime import datetime

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ==========================================
# SHEET 1: Use Case Inventory
# ==========================================
ws1 = wb.create_sheet("1-Use-Case-Inventory", 0)

# Headers
headers1 = ['ID', 'Name', 'Summary', 'Category', 'Strategic Theme', 'Owner/Architect', 'AI Type', 'Value Description']
ws1.append(headers1)

# Style headers
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Sample data (24 use cases from BNZ)
sample_data1 = [
    ['UC-001', 'FrontLine - Partnership Banking', 'GenAI-driven RM solutions for deepening banker-customer relationships', 'Customer & Relationship Management', 'Customer Experience', 'Shawn Cawood', 'Embedded/Vendor AI', 'Increased revenue through improved cross-sell/upsell effectiveness'],
    ['UC-002', 'Finance', 'Transform raw financial data into well-structured reports (P&L, balance sheets)', 'Finance & Accounting', 'Operational Efficiency', 'Brian Stapleton', 'Generative AI', 'Reduced month-end close time, improved accuracy'],
    ['UC-003', 'Analytics and Reporting', 'AI builds on BI by turning historical data into proactive guidance', 'Data & Analytics', 'Data-Driven Decisions', 'Tracey Davis', 'Generative AI', 'Democratized data access reducing analyst dependency'],
    ['UC-004', 'Credit Risk', 'GenAI-driven Credit Risk solutions for data-driven decision-making', 'Risk Management', 'Risk & Compliance', 'Paul Furkert', 'ML/AI', 'Reduced credit losses through better risk assessment'],
    ['UC-005', 'Lending Ops', 'AI Document Processing for Loan Applications & Compliance', 'Lending & Credit Operations', 'Operational Efficiency', 'Paul Furkert', 'Document AI', 'Processing time reduction for loan applications'],
    ['UC-006', 'HyperPersonalization', 'Generative AI enables new customer relevance through marketing reinvention', 'Marketing & Personalization', 'Customer Experience', 'Corina Elama', 'Generative AI', 'Increased campaign conversion rates, higher engagement'],
    ['UC-007', 'Contact Centre', 'Generate call summaries, agent-assist features, customer self-service', 'Customer Service & Support', 'Customer Experience', 'Dan Barratt', 'Embedded/Vendor AI', 'Reduced average handle time, improved CSAT'],
    ['UC-008', 'Security AI', 'Use of AI to identify threats earlier with higher accuracy', 'Cybersecurity & Threat Detection', 'Risk & Compliance', 'Francis Kataino', 'ML/AI', 'Reduced threat detection time, fewer false positives'],
    ['UC-009', 'Data Products', 'Extend GenAI intervention to accelerate data product creation', 'Data & Analytics', 'Data-Driven Decisions', 'Tracey Davis', 'Generative AI', 'Accelerated data product delivery, improved quality'],
    ['UC-010', 'SDLC', 'GenAI can automate key SDLC tasks, enhancing productivity', 'Software Development & Engineering', 'Operational Efficiency', 'Glenn Bellam', 'Code AI', 'Developer productivity increased 20-30%'],
    ['UC-011', 'Fincrime', 'Name screening, OCDD, ECDD and TM narrative writing', 'Compliance & Financial Crime', 'Risk & Compliance', 'Michael Lomas', 'ML/AI', 'Reduced investigation time, improved detection accuracy'],
    ['UC-012', 'QA/QC', 'Automate data validation, detecting errors, ensuring compliance', 'Quality Assurance & Compliance', 'Operational Efficiency', 'Brian Stapleton', 'ML/AI', 'Reduced manual QA effort, improved data quality'],
    ['UC-013', 'Fraud Ops', 'AI support for case officers through pattern detection and evidence gathering', 'Fraud Detection & Prevention', 'Risk & Compliance', 'Michael Lomas', 'ML/AI', 'Reduced fraud investigation time, prevented losses'],
    ['UC-014', 'Onboarding', 'AI streamlines onboarding by automating ID verification, KYC checks', 'Customer Onboarding & KYC', 'Customer Experience', 'Brett Allport', 'Document AI', 'Reduced onboarding time from days to hours'],
    ['UC-015', 'Risk Functions', 'AI reduces Operational Risk by identifying anomalies', 'Risk Management', 'Risk & Compliance', 'TBD', 'ML/AI', 'Improved risk detection, faster assessment'],
    ['UC-016', 'IT Ops', 'Evolution of Ticketing Process and Smart Routing', 'IT Operations & Service Management', 'Operational Efficiency', 'John Marshall', 'AIOps', 'Reduced MTTR, improved SLA compliance'],
    ['UC-017', 'FrontLine - CIB', 'GenAI-driven RM solutions for Corporate and Investment Banking', 'Customer & Relationship Management', 'Customer Experience', 'Shawn Cawood', 'Generative AI', 'Increased wallet share with corporate clients'],
    ['UC-018', 'Procurement', 'Multi-Agent System for End-to-end Procurement process optimization', 'Procurement & Supply Chain', 'Operational Efficiency', 'Brian Stapleton', 'Multi-Agent AI', 'Cost savings through optimized sourcing'],
    ['UC-019', 'Payment Disputes', 'Automate end-to-end lifecycle of card payment disputes', 'Payments & Dispute Resolution', 'Customer Experience', 'TBD', 'ML/AI', 'Reduced dispute resolution time, improved win rates'],
    ['UC-020', 'Controls Testing', 'Automate IT and business control testing', 'Audit & Controls Testing', 'Risk & Compliance', 'TBD', 'ML/AI', 'Reduced testing effort, continuous monitoring'],
    ['UC-021', 'Wholesale Underwriting', 'End-to-end transformation of CIB Risk Underwriting process', 'Lending & Credit Operations', 'Operational Efficiency', 'Shawn Cawood', 'ML/AI', 'Reduced underwriting time, improved pricing'],
    ['UC-022', 'Learning Content', 'AI accelerates content creation by generating lessons and quizzes', 'Learning & Development', 'Operational Efficiency', 'Hugh Walcott', 'Generative AI', 'Reduced content creation time, improved training'],
    ['UC-023', 'Collection Management', 'AI optimises collections by predicting payment likelihood', 'Collections & Recovery', 'Operational Efficiency', 'Paul Furkert', 'ML/AI', 'Improved recovery rates, reduced collection costs'],
    ['UC-024', 'Front-end App Personalisation', 'Use of GenAI in mobile app for hyper personalisation', 'Marketing & Personalization', 'Customer Experience', 'Corina Elama', 'Generative AI', 'Increased app engagement, higher product adoption'],
]

for row_data in sample_data1:
    ws1.append(row_data)

# Set column widths
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 60
ws1.column_dimensions['D'].width = 30
ws1.column_dimensions['E'].width = 25
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 20
ws1.column_dimensions['H'].width = 50

# Freeze first row
ws1.freeze_panes = 'A2'

# ==========================================
# SHEET 2: DVF Scoring
# ==========================================
ws2 = wb.create_sheet("2-DVF-Filtering", 1)

headers2 = ['ID', 'Name', 'Desirability (1-5)', 'Viability (1-5)', 'Feasibility (1-5)', 'DVF Total', 'DVF Status', 'Notes']
ws2.append(headers2)

# Style headers
for col in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Pull IDs and Names from Sheet 1
for i in range(2, 26):  # 24 use cases
    ws2.cell(row=i, column=1, value=f"='1-Use-Case-Inventory'!A{i}")
    ws2.cell(row=i, column=2, value=f"='1-Use-Case-Inventory'!B{i}")

    # DVF Total formula
    ws2.cell(row=i, column=6, value=f"=C{i}+D{i}+E{i}")

    # DVF Status formula
    ws2.cell(row=i, column=7, value=f'=IF(AND(C{i}>=3,D{i}>=3,E{i}>=3),"Pass","Fail")')

# Conditional formatting for DVF Status
pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

ws2.conditional_formatting.add('G2:G25',
    CellIsRule(operator='equal', formula=['"Pass"'], fill=pass_fill))
ws2.conditional_formatting.add('G2:G25',
    CellIsRule(operator='equal', formula=['"Fail"'], fill=fail_fill))

# Set column widths
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 12
ws2.column_dimensions['H'].width = 40

ws2.freeze_panes = 'A2'

# Add instructions at top
ws2.insert_rows(1)
ws2['A1'] = 'INSTRUCTIONS: Score each dimension 1-5. Pass requires ALL dimensions ≥3. 5=Excellent, 4=Good, 3=Acceptable, 2=Poor, 1=Very Poor'
ws2.merge_cells('A1:H1')
ws2['A1'].font = Font(bold=True, italic=True, color='0070C0')
ws2['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[1].height = 30

# ==========================================
# SHEET 3: WSJF Scoring
# ==========================================
ws3 = wb.create_sheet("3-WSJF-Scoring", 2)

headers3 = ['ID', 'Name', 'User/Business Value', 'Time Criticality', 'Risk Reduction/Opp Enable', 'Cost of Delay', 'Job Size', 'WSJF Score', 'WSJF Rank']
ws3.append(headers3)

# Style headers
for col in range(1, len(headers3) + 1):
    cell = ws3.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Pull IDs and Names from Sheet 1, but only if DVF Pass
for i in range(2, 26):  # 24 use cases
    ws3.cell(row=i, column=1, value=f"='1-Use-Case-Inventory'!A{i}")
    ws3.cell(row=i, column=2, value=f"='1-Use-Case-Inventory'!B{i}")

    # Cost of Delay formula
    ws3.cell(row=i, column=6, value=f"=C{i}+D{i}+E{i}")

    # WSJF Score formula
    ws3.cell(row=i, column=8, value=f"=IF(G{i}>0,F{i}/G{i},0)")
    ws3.cell(row=i, column=8).number_format = '0.00'

    # WSJF Rank formula
    ws3.cell(row=i, column=9, value=f"=IF(H{i}>0,RANK(H{i},H$2:H$25,0),\"\")")

# Conditional formatting for WSJF Rank
top5_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
top10_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
top15_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

ws3.conditional_formatting.add('I2:I25',
    CellIsRule(operator='lessThanOrEqual', formula=['5'], fill=top5_fill))
ws3.conditional_formatting.add('I2:I25',
    CellIsRule(operator='between', formula=['6', '10'], fill=top10_fill))
ws3.conditional_formatting.add('I2:I25',
    CellIsRule(operator='between', formula=['11', '15'], fill=top15_fill))

# Set column widths
ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 10
ws3.column_dimensions['H'].width = 12
ws3.column_dimensions['I'].width = 12

ws3.freeze_panes = 'A2'

# Add instructions
ws3.insert_rows(1)
ws3['A1'] = 'INSTRUCTIONS: Use Fibonacci (1,2,3,5,8,13,20). WSJF = Cost of Delay ÷ Job Size. Higher WSJF = Higher Priority.'
ws3.merge_cells('A1:I1')
ws3['A1'].font = Font(bold=True, italic=True, color='0070C0')
ws3['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws3.row_dimensions[1].height = 30

# ==========================================
# SHEET 4: Dependencies & Scheduling
# ==========================================
ws4 = wb.create_sheet("4-Dependencies-Schedule", 3)

headers4 = ['ID', 'Name', 'WSJF Rank', 'Prerequisites (IDs)', 'Enables (IDs)', 'Earliest Start', 'Recommended Wave', 'Rationale']
ws4.append(headers4)

# Style headers
for col in range(1, len(headers4) + 1):
    cell = ws4.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Pull data from previous sheets
for i in range(2, 26):
    ws4.cell(row=i, column=1, value=f"='1-Use-Case-Inventory'!A{i}")
    ws4.cell(row=i, column=2, value=f"='1-Use-Case-Inventory'!B{i}")
    ws4.cell(row=i, column=3, value=f"='3-WSJF-Scoring'!I{i+1}")  # +1 because of instruction row

# Set column widths
ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 30
ws4.column_dimensions['E'].width = 30
ws4.column_dimensions['F'].width = 15
ws4.column_dimensions['G'].width = 18
ws4.column_dimensions['H'].width = 50

ws4.freeze_panes = 'A2'

# Add instructions
ws4.insert_rows(1)
ws4['A1'] = 'INSTRUCTIONS: Document dependencies. Assign waves: Wave 1 (Q1-Q3 2026), Wave 2 (Q4 2026-Q3 2027), Wave 3 (Q4 2027-Q2 2028)'
ws4.merge_cells('A1:H1')
ws4['A1'].font = Font(bold=True, italic=True, color='0070C0')
ws4['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws4.row_dimensions[1].height = 30

# ==========================================
# SHEET 5: Portfolio View
# ==========================================
ws5 = wb.create_sheet("5-Portfolio-View", 4)

headers5 = ['ID', 'Name', 'WSJF Rank', 'Cost of Delay', 'Job Size', 'Value Tier', 'Effort Tier', 'Quadrant', 'Wave', 'Top 10', 'Top 5', 'Executive Notes']
ws5.append(headers5)

# Style headers
for col in range(1, len(headers5) + 1):
    cell = ws5.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Pull data from previous sheets
for i in range(2, 26):
    ws5.cell(row=i, column=1, value=f"='1-Use-Case-Inventory'!A{i}")
    ws5.cell(row=i, column=2, value=f"='1-Use-Case-Inventory'!B{i}")
    ws5.cell(row=i, column=3, value=f"='3-WSJF-Scoring'!I{i+1}")
    ws5.cell(row=i, column=4, value=f"='3-WSJF-Scoring'!F{i+1}")
    ws5.cell(row=i, column=5, value=f"='3-WSJF-Scoring'!G{i+1}")
    ws5.cell(row=i, column=9, value=f"='4-Dependencies-Schedule'!G{i+1}")

    # Value Tier formula
    ws5.cell(row=i, column=6, value=f'=IF(D{i}>=30,"High",IF(D{i}>=15,"Medium","Low"))')

    # Effort Tier formula
    ws5.cell(row=i, column=7, value=f'=IF(E{i}>=13,"High",IF(E{i}>=5,"Medium","Low"))')

    # Quadrant formula
    ws5.cell(row=i, column=8, value=f'=IF(F{i}="High",IF(G{i}="Low","Quick Win","Strategic Bet"),IF(F{i}="Medium","Fill-in","Avoid"))')

# Conditional formatting for Quadrants
quickwin_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
strategic_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
fillin_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
avoid_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

ws5.conditional_formatting.add('H2:H25',
    CellIsRule(operator='equal', formula=['"Quick Win"'], fill=quickwin_fill))
ws5.conditional_formatting.add('H2:H25',
    CellIsRule(operator='equal', formula=['"Strategic Bet"'], fill=strategic_fill))
ws5.conditional_formatting.add('H2:H25',
    CellIsRule(operator='equal', formula=['"Fill-in"'], fill=fillin_fill))
ws5.conditional_formatting.add('H2:H25',
    CellIsRule(operator='equal', formula=['"Avoid"'], fill=avoid_fill))

# Set column widths
ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 35
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 10
ws5.column_dimensions['F'].width = 12
ws5.column_dimensions['G'].width = 12
ws5.column_dimensions['H'].width = 15
ws5.column_dimensions['I'].width = 10
ws5.column_dimensions['J'].width = 10
ws5.column_dimensions['K'].width = 10
ws5.column_dimensions['L'].width = 50

ws5.freeze_panes = 'A2'

# Add instructions
ws5.insert_rows(1)
ws5['A1'] = 'INSTRUCTIONS: Portfolio view with automatic tier calculations. Mark Top 10 and Top 5 checkboxes. Add Executive Notes for presentation.'
ws5.merge_cells('A1:L1')
ws5['A1'].font = Font(bold=True, italic=True, color='0070C0')
ws5['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws5.row_dimensions[1].height = 30

# ==========================================
# SHEET 6: Scoring Guidelines
# ==========================================
ws6 = wb.create_sheet("6-Scoring-Guidelines", 5)

ws6['A1'] = 'WSJF SCORING GUIDELINES'
ws6['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws6.merge_cells('A1:D1')

guidelines_data = [
    ['', '', '', ''],
    ['COMPONENT', 'SCORE', 'DESCRIPTION', 'EXAMPLES'],
    ['', '', '', ''],
    ['User/Business Value', '20', '>$5M annual benefit OR >20% efficiency gain', 'Major revenue driver, critical customer experience'],
    ['', '13', '$2-5M benefit OR 10-20% efficiency', 'Significant cost savings, important customer feature'],
    ['', '8', '$500K-2M benefit OR 5-10% efficiency', 'Meaningful operational improvement'],
    ['', '5', '$100K-500K benefit OR 2-5% efficiency', 'Moderate business value'],
    ['', '3', '$50K-100K benefit OR 1-2% efficiency', 'Small but tangible benefit'],
    ['', '2', '<$50K benefit OR <1% efficiency', 'Minimal quantified benefit'],
    ['', '1', 'Negligible benefit', 'No clear business value'],
    ['', '', '', ''],
    ['Time Criticality', '20', 'Fixed regulatory deadline <6 months OR major competitive threat', 'Compliance deadline with penalties'],
    ['', '13', 'Fixed deadline <12 months OR significant competitive pressure', 'Market window closing'],
    ['', '8', 'Preferred timing <12 months OR moderate competitive benefit', 'Seasonal opportunity'],
    ['', '5', 'Beneficial timing <18 months', 'Loose timing preference'],
    ['', '3', 'Timing preference <24 months', 'Minor timing consideration'],
    ['', '2', 'No specific timing driver', 'Flexible timing'],
    ['', '1', 'Can be done anytime', 'No urgency'],
    ['', '', '', ''],
    ['Risk Reduction/', '20', 'Eliminates critical risk OR enables 5+ use cases', 'Major security vulnerability, platform foundation'],
    ['Opportunity Enablement', '13', 'Significantly reduces high risk OR enables 3-4 use cases', 'Compliance risk, key infrastructure'],
    ['', '8', 'Reduces moderate risk OR enables 2 use cases', 'Technical debt, reusable component'],
    ['', '5', 'Reduces minor risk OR enables 1 use case', 'Small risk mitigation'],
    ['', '3', 'Small risk reduction OR creates reusable component', 'Incremental improvement'],
    ['', '2', 'Minimal risk impact', 'Negligible risk reduction'],
    ['', '1', 'No risk reduction or enabling capability', 'No strategic value'],
    ['', '', '', ''],
    ['Job Size', '20', '18+ months OR >100 person-months', 'Enterprise transformation'],
    ['', '13', '12-18 months OR 50-100 person-months', 'Major program'],
    ['', '8', '6-12 months OR 20-50 person-months', 'Significant project'],
    ['', '5', '3-6 months OR 10-20 person-months', 'Standard project'],
    ['', '3', '1-3 months OR 5-10 person-months', 'Small project'],
    ['', '2', '<1 month OR <5 person-months', 'Very small initiative'],
    ['', '1', 'Minimal effort', 'Trivial change'],
]

for row in guidelines_data:
    ws6.append(row)

# Style guidelines
for row in ws6.iter_rows(min_row=2, max_row=2, min_col=1, max_col=4):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
ws6.column_dimensions['A'].width = 25
ws6.column_dimensions['B'].width = 10
ws6.column_dimensions['C'].width = 50
ws6.column_dimensions['D'].width = 40

# Add DVF Guidelines
ws6['A40'] = 'DVF SCORING GUIDELINES'
ws6['A40'].font = Font(bold=True, size=14, color='1F4E78')
ws6.merge_cells('A40:D40')

dvf_guidelines = [
    ['', '', '', ''],
    ['DIMENSION', 'SCORE', 'DESCRIPTION', 'KEY QUESTIONS'],
    ['', '', '', ''],
    ['Desirability', '5', 'Excellent - Strong evidence users/customers want this', 'Does it solve a critical pain point? Is demand validated?'],
    ['', '4', 'Good - Clear user demand', 'Is there user research supporting this?'],
    ['', '3', 'Acceptable - Meets minimum user need', 'Do users see value in this?'],
    ['', '2', 'Poor - Unclear if users want this', 'Is this an assumption or validated need?'],
    ['', '1', 'Very Poor - Users unlikely to adopt', 'Are we solving the wrong problem?'],
    ['', '', '', ''],
    ['Viability', '5', 'Excellent - Strong financial case, strategic alignment', 'Does it align with strategy? Is ROI clear?'],
    ['', '4', 'Good - Clear business case', 'Can we afford it? Does it fit our model?'],
    ['', '3', 'Acceptable - Financially viable', 'Is the business case sound?'],
    ['', '2', 'Poor - Weak business case', 'Are the financials uncertain?'],
    ['', '1', 'Very Poor - Not financially viable', 'Can we justify the investment?'],
    ['', '', '', ''],
    ['Feasibility', '5', 'Excellent - All capabilities exist, low technical risk', 'Do we have the technology, data, and skills?'],
    ['', '4', 'Good - Minor gaps, manageable risk', 'Are there any technical blockers?'],
    ['', '3', 'Acceptable - Can be built with moderate effort', 'Is it technically possible with current capabilities?'],
    ['', '2', 'Poor - Significant technical challenges', 'Do we need new technology or data?'],
    ['', '1', 'Very Poor - Major technical blockers', 'Is this even possible with current tech?'],
]

row_start = 41
for row in dvf_guidelines:
    ws6.append(row)

# Style DVF headers
for row in ws6.iter_rows(min_row=41, max_row=41, min_col=1, max_col=4):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

# ==========================================
# SHEET 7: Dashboard Summary
# ==========================================
ws7 = wb.create_sheet("7-Dashboard", 6)

ws7['A1'] = 'BNZ AI USE CASE PRIORITIZATION DASHBOARD'
ws7['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws7.merge_cells('A1:F1')
ws7['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws7.row_dimensions[1].height = 25

# Add summary statistics
ws7['A3'] = 'SUMMARY STATISTICS'
ws7['A3'].font = Font(bold=True, size=12)
ws7.merge_cells('A3:B3')

summary_labels = [
    'Total Use Cases:',
    'DVF Pass:',
    'DVF Fail:',
    'Top 10 Count:',
    'Top 5 Count:',
    'Wave 1 Use Cases:',
    'Wave 2 Use Cases:',
    'Wave 3 Use Cases:',
]

row_num = 4
for label in summary_labels:
    ws7[f'A{row_num}'] = label
    ws7[f'A{row_num}'].font = Font(bold=True)
    row_num += 1

# Add formulas for counts
ws7['B4'] = "=COUNTA('1-Use-Case-Inventory'!A2:A25)"
ws7['B5'] = '=COUNTIF(\'2-DVF-Filtering\'!G2:G25,"Pass")'
ws7['B6'] = '=COUNTIF(\'2-DVF-Filtering\'!G2:G25,"Fail")'
ws7['B7'] = "=COUNTIF('5-Portfolio-View'!J2:J25,TRUE)"
ws7['B8'] = "=COUNTIF('5-Portfolio-View'!K2:K25,TRUE)"
ws7['B9'] = '=COUNTIF(\'4-Dependencies-Schedule\'!G2:G25,"Wave 1")'
ws7['B10'] = '=COUNTIF(\'4-Dependencies-Schedule\'!G2:G25,"Wave 2")'
ws7['B11'] = '=COUNTIF(\'4-Dependencies-Schedule\'!G2:G25,"Wave 3")'

# Quadrant distribution
ws7['A13'] = 'PORTFOLIO QUADRANTS'
ws7['A13'].font = Font(bold=True, size=12)
ws7.merge_cells('A13:B13')

quadrant_labels = ['Quick Wins:', 'Strategic Bets:', 'Fill-ins:', 'Avoid:']
row_num = 14
for label in quadrant_labels:
    ws7[f'A{row_num}'] = label
    ws7[f'A{row_num}'].font = Font(bold=True)
    row_num += 1

ws7['B14'] = '=COUNTIF(\'5-Portfolio-View\'!H2:H25,"Quick Win")'
ws7['B15'] = '=COUNTIF(\'5-Portfolio-View\'!H2:H25,"Strategic Bet")'
ws7['B16'] = '=COUNTIF(\'5-Portfolio-View\'!H2:H25,"Fill-in")'
ws7['B17'] = '=COUNTIF(\'5-Portfolio-View\'!H2:H25,"Avoid")'

# Top 10 List
ws7['D3'] = 'TOP 10 USE CASES (BY WSJF RANK)'
ws7['D3'].font = Font(bold=True, size=12)
ws7.merge_cells('D3:F3')

ws7['D4'] = 'Rank'
ws7['E4'] = 'ID'
ws7['F4'] = 'Name'
for col in ['D', 'E', 'F']:
    ws7[f'{col}4'].fill = header_fill
    ws7[f'{col}4'].font = header_font

# Note: Top 10 would need to be populated with a more complex formula or manual entry
# For now, adding placeholder
for i in range(5, 15):
    ws7[f'D{i}'] = f'={i-4}'

ws7.column_dimensions['A'].width = 25
ws7.column_dimensions['B'].width = 15
ws7.column_dimensions['C'].width = 5
ws7.column_dimensions['D'].width = 10
ws7.column_dimensions['E'].width = 12
ws7.column_dimensions['F'].width = 35

# Save workbook
output_path = 'd:/Work/BNZ/ai-platform-architecture/01-motivation/03-use-cases/use-case-prioritisation/BNZ-Use-Case-Prioritization.xlsx'
wb.save(output_path)

print(f"Spreadsheet created successfully: {output_path}")
print(f"Total sheets created: {len(wb.sheetnames)}")
print(f"Sheets: {', '.join(wb.sheetnames)}")
