#!/usr/bin/env python3
"""
Script: generate-abb-catalog-excel.py
Purpose: Generate an Excel spreadsheet with all ABBs including ID, name, description, technologies, use cases, and patterns
Author: BNZ Enterprise Architecture
Date: 2025-12-06
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Configuration
OUTPUT_FILE = Path(r"D:\Work\BNZ\ai-platform-architecture\03-building-blocks\architecture-building-blocks\ABB-Catalog.xlsx")

# Comprehensive ABB catalog with technologies, use cases, patterns, and architecture layer
# Format: (ABB_ID, ABB_Name, Category, Description, Technologies, Use_Cases, Patterns, Architecture_Layer)
# Architecture Layers: AI/ML, Data & Knowledge, Integration, Infrastructure, Security, Governance, Observability
ABBS = [
    # GOV (Governance) Category
    ("AB-060", "AI Model Registry", "Governance",
     "Central registry for AI/ML models with metadata, versioning, and lifecycle management",
     "MLflow, Amazon SageMaker Model Registry, Weights & Biases",
     "UC-001, UC-004, UC-005, UC-011, UC-013, UC-021",
     "PT-001, PT-002", "AI/ML"),
    ("AB-061", "Risk Assessment Engine", "Governance",
     "Automated risk evaluation for AI models and use cases",
     "Custom Python, AWS Lambda, Risk scoring frameworks",
     "UC-004, UC-011, UC-013, UC-015, UC-021",
     "PT-001", "Governance"),
    ("AB-062", "Approval Workflow Orchestrator", "Governance",
     "Manages approval workflows for model deployment and changes",
     "AWS Step Functions, Apache Airflow, ServiceNow",
     "UC-001, UC-005, UC-014, UC-021",
     "PT-001", "Governance"),
    ("AB-063", "ML Model Explainability Engine", "Governance",
     "Generates explanations for ML model predictions using SHAP/LIME",
     "SHAP, LIME, AWS SageMaker Clarify, Alibi Explain",
     "UC-004, UC-005, UC-011, UC-013, UC-021",
     "PT-001, PT-004", "AI/ML"),
    ("AB-064", "Compliance Dashboard", "Governance",
     "Dashboard for tracking AI governance and compliance metrics",
     "Grafana, Amazon QuickSight, Tableau",
     "UC-011, UC-015, UC-020",
     "PT-001", "Governance"),
    ("AB-065", "Audit Trail & Logging", "Governance",
     "Comprehensive audit logging for AI system activities",
     "Amazon CloudWatch, OpenSearch, Splunk",
     "UC-011, UC-013, UC-015, UC-020",
     "PT-001, PT-017", "Governance"),
    ("AB-066", "AI Gateway", "Governance",
     "Central gateway for AI service access and policy enforcement",
     "Amazon API Gateway, Kong, AWS Lambda",
     "UC-001, UC-007, UC-010",
     "PT-001, PT-006", "Integration"),
    ("AB-067", "Cost Attribution Engine", "Governance",
     "Tracks and attributes AI infrastructure costs to use cases",
     "AWS Cost Explorer, CloudHealth, custom tagging",
     "UC-002, UC-003",
     "PT-001", "Governance"),
    ("AB-068", "Explainability Service", "Governance",
     "Service for providing model explanations to stakeholders",
     "SHAP, LIME, custom explanation APIs",
     "UC-004, UC-005, UC-021",
     "PT-004", "AI/ML"),
    ("AB-069", "Bias Detection Framework", "Governance",
     "Detects and reports bias in ML models across demographic groups",
     "AWS SageMaker Clarify, Fairlearn, AI Fairness 360",
     "UC-004, UC-005, UC-006, UC-014",
     "PT-001, PT-004", "AI/ML"),
    ("AB-070", "Data Lineage Tracker", "Governance",
     "Tracks data provenance and lineage for AI models",
     "Apache Atlas, AWS Glue Data Catalog, OpenLineage",
     "UC-003, UC-009, UC-012",
     "PT-001, PT-012", "Data & Knowledge"),
    ("AB-071", "Model Performance Benchmarking", "Governance",
     "Benchmarks model performance against baselines and competitors",
     "MLflow, custom benchmarking frameworks",
     "UC-004, UC-005, UC-021",
     "PT-001, PT-002", "AI/ML"),

    # SEC (Security) Category
    ("AB-112", "Data Encryption Service", "Security",
     "Encryption services for data at rest and in transit",
     "AWS KMS, AWS Certificate Manager, TLS 1.3",
     "All use cases",
     "PT-018", "Security"),
    ("AB-113", "Access Control & Identity Management", "Security",
     "Identity and access management for AI platform",
     "AWS IAM, AWS IAM Identity Center, Okta",
     "All use cases",
     "PT-018", "Security"),
    ("AB-114", "PII Detection & Masking Service", "Security",
     "Detects and masks personally identifiable information",
     "AWS Macie, Amazon Comprehend PII, Presidio",
     "UC-006, UC-007, UC-014, UC-019",
     "PT-018", "Security"),
    ("AB-115", "Model Security Framework", "Security",
     "Security controls and vulnerability management for models",
     "Model scanning tools, adversarial testing frameworks",
     "UC-008, UC-011, UC-013",
     "PT-018", "Security"),
    ("AB-116", "Audit Logging & Monitoring", "Security",
     "Security-focused audit logging and threat monitoring",
     "AWS CloudTrail, AWS Security Hub, SIEM integration",
     "All use cases",
     "PT-017, PT-018", "Security"),
    ("AB-117", "Privacy Compliance Engine", "Security",
     "Ensures compliance with privacy regulations (GDPR, Privacy Act)",
     "Custom compliance checks, consent management",
     "UC-006, UC-007, UC-014",
     "PT-018", "Security"),
    ("AB-118", "Secrets Management", "Security",
     "Secure storage and management of credentials and secrets",
     "AWS Secrets Manager, HashiCorp Vault",
     "All use cases",
     "PT-018", "Security"),
    ("AB-119", "Differential Privacy Engine", "Security",
     "Implements differential privacy for sensitive data analytics",
     "OpenDP, Google DP library, PySyft",
     "UC-003, UC-006",
     "PT-018", "Security"),
    ("AB-120", "Federated Learning Platform", "Security",
     "Enables privacy-preserving distributed model training",
     "PySyft, TensorFlow Federated, NVIDIA FLARE",
     "UC-004, UC-011",
     "PT-018", "Security"),
    ("AB-121", "Homomorphic Encryption Service", "Security",
     "Computation on encrypted data without decryption",
     "Microsoft SEAL, IBM HELib, Concrete",
     "UC-004, UC-011",
     "PT-018", "Security"),
    ("AB-122", "Secure Multi-Party Computation", "Security",
     "Enables secure computation across multiple parties",
     "MP-SPDZ, Sharemind, custom MPC implementations",
     "UC-004, UC-011",
     "PT-018", "Security"),
    ("AB-123", "Data Lineage & Provenance Tracker", "Security",
     "Security-focused data lineage and provenance tracking",
     "Apache Atlas, OpenLineage, AWS Glue",
     "UC-009, UC-011, UC-012",
     "PT-012, PT-018", "Data & Knowledge"),
    ("AB-124", "Model Watermarking Service", "Security",
     "Embeds watermarks in models for IP protection",
     "Custom watermarking algorithms, model fingerprinting",
     "UC-010",
     "PT-018", "Security"),
    ("AB-125", "Adversarial Testing Framework", "Security",
     "Tests models against adversarial attacks",
     "Adversarial Robustness Toolbox, CleverHans, Foolbox",
     "UC-008, UC-011, UC-013",
     "PT-018", "Security"),

    # OBS (Observability) Category
    ("AB-096", "Observability Platform", "Observability",
     "Centralized platform for metrics, logs, and traces",
     "Amazon CloudWatch, Datadog, Grafana, Prometheus",
     "All use cases",
     "PT-017", "Observability"),
    ("AB-097", "Data Drift Detector", "Observability",
     "Detects drift in input data distributions",
     "Evidently AI, NannyML, AWS SageMaker Model Monitor",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-017", "Observability"),
    ("AB-098", "Prediction Drift Monitor", "Observability",
     "Monitors drift in model predictions over time",
     "Evidently AI, NannyML, custom monitoring",
     "UC-004, UC-005, UC-013",
     "PT-017", "Observability"),
    ("AB-099", "Concept Drift Detector", "Observability",
     "Detects changes in underlying data relationships",
     "Evidently AI, River, custom drift detection",
     "UC-004, UC-005, UC-013",
     "PT-017", "Observability"),
    ("AB-100", "Cost Monitoring Dashboard", "Observability",
     "Real-time monitoring of AI infrastructure costs",
     "AWS Cost Explorer, Kubecost, custom dashboards",
     "UC-002, UC-003",
     "PT-001, PT-017", "Observability"),
    ("AB-101", "Latency Monitor", "Observability",
     "Tracks inference and pipeline latency metrics",
     "Amazon CloudWatch, Datadog APM, custom metrics",
     "UC-007, UC-009",
     "PT-009, PT-017", "Observability"),
    ("AB-102", "Error Tracking System", "Observability",
     "Captures and categorizes errors across AI systems",
     "Sentry, AWS CloudWatch, PagerDuty",
     "All use cases",
     "PT-017", "Observability"),
    ("AB-103", "Business Metrics Tracker", "Observability",
     "Links model performance to business outcomes",
     "Custom metrics, Amazon QuickSight, Looker",
     "UC-001, UC-002, UC-006",
     "PT-017", "Observability"),
    ("AB-104", "Data Quality Monitor", "Observability",
     "Monitors data quality dimensions in real-time",
     "Great Expectations, Deequ, AWS Glue Data Quality",
     "UC-003, UC-009, UC-012",
     "PT-003, PT-017", "Observability"),
    ("AB-105", "GenAI Quality Monitor", "Observability",
     "Monitors quality of generative AI outputs",
     "LangSmith, Weights & Biases, custom evaluation",
     "UC-001, UC-007, UC-010, UC-022",
     "PT-005, PT-017", "Observability"),
    ("AB-106", "Automated Retraining Trigger", "Observability",
     "Triggers model retraining based on drift thresholds",
     "AWS Step Functions, Apache Airflow, custom triggers",
     "UC-004, UC-005, UC-013",
     "PT-002, PT-017", "Observability"),
    ("AB-107", "Stakeholder Dashboard", "Observability",
     "Executive dashboards for AI platform visibility",
     "Amazon QuickSight, Tableau, Grafana",
     "UC-002, UC-003, UC-015",
     "PT-017", "Observability"),

    # INT (Integration) Category
    ("AB-074", "Event Broker", "Integration",
     "Central event broker for async communication",
     "Amazon Kinesis, Apache Kafka, Amazon MSK",
     "UC-008, UC-013, UC-016, UC-019",
     "PT-015", "Integration"),
    ("AB-075", "Stream Processing Engine", "Integration",
     "Processes streaming data in real-time",
     "Amazon Kinesis Analytics, Apache Flink, Spark Streaming",
     "UC-008, UC-013, UC-016",
     "PT-016", "Integration"),
    ("AB-076", "Event Producer Adapter", "Integration",
     "Adapters for producing events from various sources",
     "AWS Lambda, Kafka Connect, custom producers",
     "UC-008, UC-013, UC-016",
     "PT-015", "Integration"),
    ("AB-077", "Event Consumer Framework", "Integration",
     "Framework for consuming and processing events",
     "AWS Lambda, Kafka Consumers, custom consumers",
     "UC-008, UC-013, UC-016",
     "PT-015", "Integration"),
    ("AB-078", "Dead Letter Queue", "Integration",
     "Handles failed message processing",
     "Amazon SQS DLQ, Kafka DLQ topics",
     "UC-008, UC-013, UC-019",
     "PT-015", "Integration"),
    ("AB-079", "Event Replay Service", "Integration",
     "Replays events for recovery and testing",
     "Custom replay service, Kafka offset management",
     "UC-008, UC-013",
     "PT-015", "Integration"),

    # GEN (Generative AI) Category
    ("AB-050", "Large Language Model Service", "GenAI",
     "Foundation LLM service for text generation",
     "AWS Bedrock (Claude, Llama), Amazon SageMaker JumpStart",
     "UC-001, UC-007, UC-010, UC-017, UC-022",
     "PT-005, PT-006, PT-007", "AI/ML"),
    ("AB-051", "Vector Database", "GenAI",
     "Vector storage for embeddings and similarity search",
     "Amazon OpenSearch, Pinecone, Weaviate, pgvector",
     "UC-001, UC-005, UC-007, UC-010",
     "PT-005", "Data & Knowledge"),
    ("AB-052", "Semantic Search Engine", "GenAI",
     "Semantic search across knowledge bases",
     "Amazon OpenSearch, Amazon Kendra, custom embeddings",
     "UC-001, UC-007, UC-010",
     "PT-005", "AI/ML"),
    ("AB-053", "Query Intent Analyzer", "GenAI",
     "Analyzes and classifies query intent",
     "AWS Bedrock, Amazon Comprehend, custom classifiers",
     "UC-001, UC-007",
     "PT-005, PT-013", "AI/ML"),
    ("AB-054", "Hybrid Search Engine", "GenAI",
     "Combines keyword and semantic search",
     "Amazon OpenSearch, Elasticsearch, custom fusion",
     "UC-001, UC-007, UC-010",
     "PT-005", "AI/ML"),
    ("AB-055", "Reranking Engine", "GenAI",
     "Reranks search results for relevance",
     "Cohere Rerank, cross-encoder models, custom rerankers",
     "UC-001, UC-007",
     "PT-005", "AI/ML"),
    ("AB-056", "Self-Critique Engine", "GenAI",
     "LLM self-evaluation and refinement",
     "AWS Bedrock, custom prompt chains",
     "UC-001, UC-007, UC-010",
     "PT-005, PT-007", "AI/ML"),
    ("AB-057", "Document Processing Pipeline", "GenAI",
     "Processes documents for RAG ingestion",
     "LangChain, LlamaIndex, custom chunking",
     "UC-001, UC-005, UC-011, UC-014",
     "PT-005, PT-011", "Data & Knowledge"),
    ("AB-058", "Citation Generator", "GenAI",
     "Generates citations for LLM responses",
     "Custom citation logic, LangChain",
     "UC-001, UC-007, UC-010",
     "PT-005", "AI/ML"),
    ("AB-059", "Query Rewriting Engine", "GenAI",
     "Rewrites queries for better retrieval",
     "AWS Bedrock, custom query expansion",
     "UC-001, UC-007",
     "PT-005", "AI/ML"),

    # AGT (Agentic AI) Category
    ("AB-001", "Agent Orchestrator", "Agentic",
     "Coordinates agent collaboration and task execution",
     "LangGraph, AutoGen, CrewAI, AWS Bedrock Agents",
     "UC-001, UC-007, UC-010, UC-016, UC-017",
     "PT-007, PT-008", "AI/ML"),
    ("AB-002", "Tool Registry", "Agentic",
     "Registry of tools available to AI agents",
     "Custom tool registry, LangChain tools",
     "UC-001, UC-010, UC-016",
     "PT-007", "AI/ML"),
    ("AB-003", "Agent Memory System", "Agentic",
     "Persistent memory for agent context",
     "Redis, Amazon DynamoDB, vector stores",
     "UC-001, UC-007, UC-010",
     "PT-007", "Data & Knowledge"),
    ("AB-004", "Planning Module", "Agentic",
     "Strategic planning for agent task execution",
     "LangGraph, custom planning algorithms",
     "UC-001, UC-010, UC-016",
     "PT-007", "AI/ML"),
    ("AB-005", "Execution Engine", "Agentic",
     "Executes agent plans and actions",
     "AWS Lambda, custom execution runtime",
     "UC-001, UC-010, UC-016",
     "PT-007", "AI/ML"),
    ("AB-006", "Reflection Module", "Agentic",
     "Agent self-reflection and learning",
     "LangGraph, custom reflection prompts",
     "UC-001, UC-010",
     "PT-007", "AI/ML"),
    ("AB-007", "Multi-Agent Communication Bus", "Agentic",
     "Inter-agent communication infrastructure",
     "Amazon SQS, Redis Pub/Sub, custom messaging",
     "UC-010, UC-016",
     "PT-008", "Integration"),
    ("AB-008", "Human-in-the-Loop Gateway", "Agentic",
     "Human oversight interface for agent actions",
     "Custom approval workflows, Slack integration",
     "UC-001, UC-005, UC-021",
     "PT-007", "AI/ML"),
    ("AB-009", "Agent Cost Monitor", "Agentic",
     "Monitors token and API costs for agents",
     "Custom cost tracking, LangSmith",
     "UC-001, UC-007, UC-010",
     "PT-007", "Observability"),
    ("AB-010", "Agent Versioning Manager", "Agentic",
     "Version control for agent configurations",
     "Git, MLflow, custom versioning",
     "UC-010",
     "PT-007", "AI/ML"),

    # DAT (Data) Category
    ("AB-037", "Feature Store", "Data",
     "Centralized repository for ML features",
     "Feast, AWS SageMaker Feature Store, Tecton",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-003", "Data & Knowledge"),
    ("AB-038", "Data Lake", "Data",
     "Scalable storage for structured and unstructured data",
     "Snowflake, Amazon S3, Delta Lake",
     "UC-003, UC-009, UC-012",
     "PT-012", "Data & Knowledge"),
    ("AB-039", "Data Versioning Service", "Data",
     "Version control for datasets",
     "DVC, Delta Lake, LakeFS",
     "UC-003, UC-009",
     "PT-003, PT-012", "Data & Knowledge"),

    # MLO (ML Operations) Category
    ("AB-085", "ML Training Pipeline", "MLOps",
     "Orchestrates model training workflows",
     "AWS SageMaker Pipelines, Kubeflow, MLflow",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-002", "AI/ML"),
    ("AB-086", "Model Registry", "MLOps",
     "Central registry for trained models",
     "MLflow, AWS SageMaker Model Registry",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-002", "AI/ML"),
    ("AB-087", "Model Testing Framework", "MLOps",
     "Comprehensive model testing and validation",
     "pytest, Great Expectations, custom testing",
     "UC-004, UC-005, UC-012",
     "PT-002", "AI/ML"),
    ("AB-088", "Model Deployment Orchestrator", "MLOps",
     "Automates model deployment workflows",
     "AWS SageMaker, Kubernetes, ArgoCD",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-002", "AI/ML"),
    ("AB-089", "Model Monitoring Platform", "MLOps",
     "Production model performance monitoring",
     "AWS SageMaker Model Monitor, Evidently AI, NannyML",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-002, PT-017", "AI/ML"),
    ("AB-090", "Model Fallback Logic", "MLOps",
     "Handles model failures with fallback strategies",
     "Custom fallback logic, circuit breakers",
     "UC-004, UC-006, UC-007",
     "PT-006, PT-009", "AI/ML"),
    ("AB-091", "Rollback Controller", "MLOps",
     "Manages model version rollbacks",
     "AWS SageMaker, custom rollback automation",
     "UC-004, UC-005",
     "PT-002, PT-010", "AI/ML"),
    ("AB-092", "A/B Test Configuration Service", "MLOps",
     "Configures and manages A/B tests for models",
     "Custom traffic splitting, AWS Lambda@Edge",
     "UC-004, UC-006",
     "PT-010", "AI/ML"),
    ("AB-093", "Feature Flag Service", "MLOps",
     "Feature flag management for ML features",
     "LaunchDarkly, AWS AppConfig, custom flags",
     "UC-006, UC-024",
     "PT-010", "AI/ML"),

    # DOC (Document Processing) Category
    ("AB-040", "Document Classification Engine", "Document",
     "Classifies documents by type and content",
     "AWS Textract, Amazon Comprehend, custom classifiers",
     "UC-005, UC-011, UC-014, UC-018",
     "PT-011", "AI/ML"),
    ("AB-041", "OCR Engine", "Document",
     "Optical character recognition for documents",
     "AWS Textract, Tesseract, custom OCR",
     "UC-005, UC-011, UC-014",
     "PT-011", "AI/ML"),
    ("AB-042", "Document Forensics Engine", "Document",
     "Detects document tampering and fraud",
     "Custom forensics algorithms, image analysis",
     "UC-011, UC-013, UC-014",
     "PT-011", "AI/ML"),

    # NLP (Natural Language Processing) Category
    ("AB-095", "NLP Extraction Engine", "NLP",
     "Extracts entities and relationships from text",
     "Amazon Comprehend, spaCy, custom NER models",
     "UC-005, UC-011, UC-014, UC-019",
     "PT-011", "AI/ML"),

    # VAL (Validation) Category
    ("AB-127", "Validation Engine", "Validation",
     "Validates data and model outputs",
     "Great Expectations, Pydantic, custom validation",
     "UC-005, UC-011, UC-012, UC-014",
     "PT-003, PT-011", "Data & Knowledge"),

    # WFL (Workflow) Category
    ("AB-129", "Human-in-the-Loop Workflow", "Workflow",
     "Manages human review workflows for AI decisions",
     "AWS Step Functions, custom workflow engine",
     "UC-005, UC-011, UC-014, UC-021",
     "PT-007, PT-011", "AI/ML"),

    # VIS (Visualization) Category
    ("AB-128", "Explanation Visualization Layer", "Visualization",
     "Visualizes model explanations for stakeholders",
     "Custom dashboards, Plotly, D3.js",
     "UC-004, UC-005, UC-021",
     "PT-004", "AI/ML"),

    # NLG (Natural Language Generation) Category
    ("AB-094", "Natural Language Generation", "NLG",
     "Generates natural language from structured data",
     "AWS Bedrock, custom templates, NLG libraries",
     "UC-001, UC-003, UC-022",
     "PT-004, PT-005", "AI/ML"),

    # FAI (Fairness) Category
    ("AB-046", "Fairness Monitoring Engine", "Fairness",
     "Monitors and reports on model fairness metrics",
     "AWS SageMaker Clarify, Fairlearn, AI Fairness 360",
     "UC-004, UC-005, UC-006, UC-014",
     "PT-001, PT-004", "AI/ML"),

    # API Category
    ("AB-011", "API Gateway", "API",
     "Central API gateway for AI services",
     "Amazon API Gateway, Kong",
     "All use cases",
     "PT-001, PT-006", "Integration"),
    ("AB-012", "Explanation API Gateway", "API",
     "API gateway for explanation services",
     "Amazon API Gateway, custom APIs",
     "UC-004, UC-005, UC-021",
     "PT-004", "Integration"),

    # CAC (Caching) Category
    ("AB-024", "Prediction Caching Layer", "Caching",
     "Caches predictions for performance optimization",
     "Amazon ElastiCache, Redis, custom caching",
     "UC-006, UC-007, UC-009",
     "PT-009", "Infrastructure"),

    # FTR (Features) Category
    ("AB-047", "Online Feature Store", "Features",
     "Real-time feature serving with <10ms latency",
     "Feast online, Redis, Amazon DynamoDB",
     "UC-004, UC-006, UC-009, UC-013",
     "PT-003, PT-009", "Data & Knowledge"),
    ("AB-048", "Feature Engineering Pipeline", "Features",
     "Automated feature engineering and transformation",
     "AWS Glue, Spark, custom pipelines",
     "UC-004, UC-005, UC-006",
     "PT-003", "Data & Knowledge"),
    ("AB-049", "Offline Feature Store", "Features",
     "Batch feature storage for model training",
     "Feast offline, Snowflake, Delta Lake",
     "UC-004, UC-005, UC-006",
     "PT-003, PT-014", "Data & Knowledge"),

    # STR (Stream) Category
    ("AB-126", "Stream Processing Platform", "Stream",
     "Platform for real-time stream processing",
     "Amazon Kinesis, Apache Flink, Spark Streaming",
     "UC-008, UC-013, UC-016",
     "PT-016", "Integration"),

    # LDB (Load Balancing) Category
    ("AB-081", "Load Balancer", "LoadBalancing",
     "Distributes load across model replicas",
     "AWS ALB, NGINX, custom load balancing",
     "UC-006, UC-007, UC-009",
     "PT-009", "Infrastructure"),

    # MDL (Model) Category
    ("AB-082", "Model Serving Infrastructure", "Model",
     "Infrastructure for serving ML models at scale",
     "AWS SageMaker Endpoints, Kubernetes, Triton",
     "UC-004, UC-005, UC-006, UC-013",
     "PT-009", "AI/ML"),
    ("AB-083", "Model Fallback Logic", "Model",
     "Fallback strategies for model failures",
     "Custom fallback, circuit breakers",
     "UC-004, UC-006, UC-007",
     "PT-006, PT-009", "AI/ML"),
    ("AB-084", "Model Quantization Service", "Model",
     "Optimizes models through quantization for faster inference",
     "ONNX, TensorRT, custom quantization",
     "UC-006, UC-007, UC-009",
     "PT-009", "AI/ML"),

    # INF (Inference) Category
    ("AB-072", "Inference Engine", "Inference",
     "Core inference execution engine for model predictions",
     "AWS SageMaker, Triton Inference Server, TorchServe",
     "UC-004, UC-005, UC-006, UC-009, UC-013",
     "PT-009", "AI/ML"),
    ("AB-073", "Inference Model Registry", "Inference",
     "Registry for inference-optimized model artifacts",
     "MLflow, AWS SageMaker Model Registry",
     "UC-004, UC-005, UC-006",
     "PT-002, PT-009", "AI/ML"),

    # BAT (Batch) Category
    ("AB-013", "Batch Orchestrator", "Batch",
     "Orchestrates batch prediction jobs and workflows",
     "AWS Step Functions, Apache Airflow, Prefect",
     "UC-004, UC-005, UC-023",
     "PT-014", "AI/ML"),
    ("AB-014", "Batch Feature Store", "Batch",
     "Feature store optimized for batch processing",
     "Feast offline, Snowflake, Delta Lake",
     "UC-004, UC-005",
     "PT-003, PT-014", "Data & Knowledge"),
    ("AB-015", "Distributed Processing Engine", "Batch",
     "Distributed compute for large-scale batch jobs",
     "Apache Spark, AWS Glue, Databricks",
     "UC-003, UC-004, UC-023",
     "PT-014", "Infrastructure"),
    ("AB-016", "Batch Model Registry", "Batch",
     "Model registry for batch prediction models",
     "MLflow, AWS SageMaker",
     "UC-004, UC-005",
     "PT-002, PT-014", "AI/ML"),
    ("AB-017", "Prediction Storage", "Batch",
     "Stores batch prediction results for downstream consumption",
     "Snowflake, Amazon S3, Delta Lake",
     "UC-004, UC-005, UC-023",
     "PT-014", "Data & Knowledge"),
    ("AB-018", "Data Quality Validator", "Batch",
     "Validates data quality in batch pipelines",
     "Great Expectations, Deequ, custom validation",
     "UC-003, UC-004, UC-012",
     "PT-003, PT-014", "Data & Knowledge"),
    ("AB-019", "Incremental Processing Manager", "Batch",
     "Manages incremental/delta batch processing",
     "Delta Lake, custom CDC logic",
     "UC-003, UC-004",
     "PT-014", "Data & Knowledge"),
    ("AB-020", "Prediction Version Tracker", "Batch",
     "Tracks versions of batch predictions over time",
     "Custom versioning, Delta Lake time travel",
     "UC-004, UC-005",
     "PT-014", "AI/ML"),
    ("AB-021", "Drift Detection Monitor", "Batch",
     "Monitors drift in batch predictions",
     "Evidently AI, custom drift detection",
     "UC-004, UC-005",
     "PT-014, PT-017", "AI/ML"),
    ("AB-022", "Result Reconciliation Engine", "Batch",
     "Reconciles batch prediction results with actuals",
     "Custom reconciliation logic, Spark",
     "UC-004, UC-005, UC-023",
     "PT-014", "AI/ML"),
    ("AB-023", "Batch Explainability Service", "Batch",
     "Generates explanations for batch predictions",
     "SHAP, custom batch explanation",
     "UC-004, UC-005",
     "PT-004, PT-014", "AI/ML"),

    # KNW (Knowledge) Category
    ("AB-080", "Knowledge Base", "Knowledge",
     "Structured knowledge storage for AI applications",
     "Amazon Kendra, vector stores, graph databases",
     "UC-001, UC-007, UC-010, UC-022",
     "PT-005", "Data & Knowledge"),

    # CNV (Conversational) Category
    ("AB-029", "Conversation State Manager", "Conversational",
     "Manages conversational context and state across turns",
     "Redis, Amazon DynamoDB, custom state management",
     "UC-007, UC-019",
     "PT-013", "AI/ML"),
    ("AB-030", "Intent Classifier", "Conversational",
     "Classifies user intent in conversations",
     "Amazon Lex, AWS Bedrock, custom classifiers",
     "UC-007, UC-019",
     "PT-013", "AI/ML"),
    ("AB-031", "Response Generator", "Conversational",
     "Generates natural language responses",
     "AWS Bedrock, Amazon Lex, custom generation",
     "UC-007, UC-019",
     "PT-013", "AI/ML"),
    ("AB-032", "Handoff Logic Engine", "Conversational",
     "Manages handoff to human agents",
     "Custom routing logic, contact center integration",
     "UC-007",
     "PT-013", "AI/ML"),
    ("AB-033", "Sentiment Analyzer", "Conversational",
     "Analyzes sentiment in conversations",
     "Amazon Comprehend, custom sentiment models",
     "UC-007, UC-019",
     "PT-013", "AI/ML"),
    ("AB-034", "Entity Extractor", "Conversational",
     "Extracts entities from conversation text",
     "Amazon Comprehend, spaCy, custom NER",
     "UC-007, UC-019",
     "PT-013", "AI/ML"),
    ("AB-035", "Conversation Summarizer", "Conversational",
     "Summarizes conversation history",
     "AWS Bedrock, custom summarization",
     "UC-007",
     "PT-013", "AI/ML"),
    ("AB-036", "Multi-Channel Orchestrator", "Conversational",
     "Orchestrates conversations across chat channels",
     "Amazon Connect, custom orchestration",
     "UC-007",
     "PT-013", "Integration"),

    # RTE (Routing) Category
    ("AB-108", "Model Router", "Routing",
     "Routes requests to appropriate models",
     "Custom routing logic, AWS Lambda",
     "UC-001, UC-007",
     "PT-006", "Integration"),
    ("AB-109", "Cost-Based Router", "Routing",
     "Routes based on cost optimization",
     "Custom cost-aware routing",
     "UC-001, UC-007, UC-010",
     "PT-006", "Integration"),
    ("AB-110", "Latency-Based Router", "Routing",
     "Routes based on latency requirements",
     "Custom latency-aware routing",
     "UC-007, UC-009",
     "PT-006", "Integration"),
    ("AB-111", "Capability-Based Router", "Routing",
     "Routes based on model capabilities",
     "Custom capability matching",
     "UC-001, UC-007, UC-010",
     "PT-006", "Integration"),

    # CMP (Champion-Challenger) Category
    ("AB-025", "Traffic Splitter", "ChampionChallenger",
     "Splits traffic between model versions for A/B testing",
     "AWS Lambda@Edge, custom traffic splitting",
     "UC-004, UC-006",
     "PT-010", "AI/ML"),
    ("AB-026", "Performance Comparator", "ChampionChallenger",
     "Compares model performance metrics",
     "Custom comparison logic, statistical libraries",
     "UC-004, UC-006",
     "PT-010", "AI/ML"),
    ("AB-027", "Statistical Test Engine", "ChampionChallenger",
     "Runs statistical significance tests",
     "SciPy, statsmodels, custom tests",
     "UC-004, UC-006",
     "PT-010", "AI/ML"),
    ("AB-028", "Model Promotion Service", "ChampionChallenger",
     "Promotes challenger to champion based on results",
     "Custom promotion logic, MLflow",
     "UC-004, UC-006",
     "PT-010", "AI/ML"),

    # DSH (Data Mesh) Category
    ("AB-043", "Data Product Catalog", "DataMesh",
     "Catalog of data products for discovery",
     "AWS Glue Data Catalog, DataHub, Atlan",
     "UC-003, UC-009",
     "PT-012", "Data & Knowledge"),
    ("AB-044", "Domain Data Platform", "DataMesh",
     "Self-serve domain data platform",
     "Snowflake, custom data platform",
     "UC-003, UC-009",
     "PT-012", "Data & Knowledge"),
    ("AB-045", "Data Contract Registry", "DataMesh",
     "Registry for data contracts between domains",
     "Custom contract registry, schema registry",
     "UC-003, UC-009",
     "PT-012", "Data & Knowledge"),
]

# Category display names and colors
CATEGORY_COLORS = {
    "Governance": "4472C4",
    "Security": "C00000",
    "Observability": "70AD47",
    "Integration": "ED7D31",
    "GenAI": "7030A0",
    "Agentic": "9933FF",
    "Data": "00B0F0",
    "MLOps": "FFC000",
    "Document": "92D050",
    "NLP": "00B050",
    "Validation": "BF8F00",
    "Workflow": "FF6600",
    "Visualization": "FF99CC",
    "NLG": "6699FF",
    "Fairness": "996633",
    "API": "333333",
    "Caching": "99CCFF",
    "Features": "66CCCC",
    "Stream": "FF9966",
    "LoadBalancing": "CCCCCC",
    "Model": "FFCC00",
    "Inference": "FF6666",
    "Batch": "9999FF",
    "Knowledge": "CC99FF",
    "Conversational": "FF66CC",
    "Routing": "66FF99",
    "ChampionChallenger": "FFFF66",
    "DataMesh": "99FF99",
}

def create_excel():
    """Create the Excel workbook with ABB catalog."""
    wb = Workbook()

    # Main ABB Catalog sheet
    ws = wb.active
    ws.title = "ABB Catalog"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["ABB ID", "Name", "Category", "Architecture Layer", "Description", "Technologies", "Use Cases", "Patterns"]
    col_widths = [15, 35, 18, 18, 60, 50, 40, 25]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Architecture layer colors
    layer_colors = {
        "AI/ML": "7030A0",           # Purple for AI/ML
        "Data & Knowledge": "00B0F0", # Blue for Data & Knowledge
        "Integration": "ED7D31",      # Orange for Integration
        "Infrastructure": "7F7F7F",   # Gray for Infrastructure
        "Security": "C00000",         # Red for Security
        "Governance": "4472C4",       # Navy for Governance
        "Observability": "FFC000",    # Gold for Observability
    }

    # Add data
    for row, abb in enumerate(ABBS, 2):
        abb_id, name, category, description, technologies, use_cases, patterns, arch_layer = abb

        data = [abb_id, name, category, arch_layer, description, technologies, use_cases, patterns]

        # Get category color
        cat_color = CATEGORY_COLORS.get(category, "FFFFFF")
        cat_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")

        # Get layer color
        layer_color = layer_colors.get(arch_layer, "FFFFFF")
        layer_fill = PatternFill(start_color=layer_color, end_color=layer_color, fill_type="solid")

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Apply category color to category column
            if col == 3:
                cell.fill = cat_fill
                if cat_color in ["C00000", "003087", "333333", "7030A0", "9933FF"]:
                    cell.font = Font(color="FFFFFF")

            # Apply layer color to architecture layer column
            if col == 4:
                cell.fill = layer_fill
                if layer_color in ["C00000", "7F7F7F", "4472C4", "7030A0"]:
                    cell.font = Font(color="FFFFFF")

    # Set row height
    for row in range(2, len(ABBS) + 2):
        ws.row_dimensions[row].height = 45

    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary by Category")

    # Count ABBs by category
    category_counts = {}
    for abb in ABBS:
        cat = abb[2]
        category_counts[cat] = category_counts.get(cat, 0) + 1

    # Summary headers
    summary_headers = ["Category", "Count", "Percentage"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 12

    # Summary data
    total = len(ABBS)
    for row, (cat, count) in enumerate(sorted(category_counts.items()), 2):
        cat_color = CATEGORY_COLORS.get(cat, "FFFFFF")
        cat_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")

        ws_summary.cell(row=row, column=1, value=cat).fill = cat_fill
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=f"{count/total*100:.1f}%")

        for col in range(1, 4):
            ws_summary.cell(row=row, column=col).border = thin_border
            if col == 1 and cat_color in ["C00000", "003087", "333333", "7030A0", "9933FF"]:
                ws_summary.cell(row=row, column=1).font = Font(color="FFFFFF")

    # Total row
    total_row = len(category_counts) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=total).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=3, value="100%").font = Font(bold=True)
    for col in range(1, 4):
        ws_summary.cell(row=total_row, column=col).border = thin_border

    # Create Summary by Architecture Layer sheet
    ws_layer = wb.create_sheet("Summary by Layer")

    # Count ABBs by architecture layer
    layer_counts = {}
    for abb in ABBS:
        layer = abb[7]  # Architecture layer is at index 7
        layer_counts[layer] = layer_counts.get(layer, 0) + 1

    # Layer summary headers
    layer_summary_headers = ["Architecture Layer", "Count", "Percentage"]
    for col, header in enumerate(layer_summary_headers, 1):
        cell = ws_layer.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws_layer.column_dimensions['A'].width = 25
    ws_layer.column_dimensions['B'].width = 12
    ws_layer.column_dimensions['C'].width = 12

    # Layer summary data
    for row, (layer, count) in enumerate(sorted(layer_counts.items()), 2):
        layer_color = layer_colors.get(layer, "FFFFFF")
        layer_fill = PatternFill(start_color=layer_color, end_color=layer_color, fill_type="solid")

        ws_layer.cell(row=row, column=1, value=layer).fill = layer_fill
        ws_layer.cell(row=row, column=2, value=count)
        ws_layer.cell(row=row, column=3, value=f"{count/total*100:.1f}%")

        for col in range(1, 4):
            ws_layer.cell(row=row, column=col).border = thin_border
            if col == 1 and layer_color in ["C00000", "7F7F7F", "4472C4", "7030A0"]:
                ws_layer.cell(row=row, column=1).font = Font(color="FFFFFF")

    # Total row
    layer_total_row = len(layer_counts) + 2
    ws_layer.cell(row=layer_total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_layer.cell(row=layer_total_row, column=2, value=total).font = Font(bold=True)
    ws_layer.cell(row=layer_total_row, column=3, value="100%").font = Font(bold=True)
    for col in range(1, 4):
        ws_layer.cell(row=layer_total_row, column=col).border = thin_border

    # Create Pattern Coverage sheet
    ws_patterns = wb.create_sheet("Pattern Coverage")

    pattern_headers = ["Pattern ID", "Pattern Name", "ABB Count", "ABBs"]
    for col, header in enumerate(pattern_headers, 1):
        cell = ws_patterns.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws_patterns.column_dimensions['A'].width = 12
    ws_patterns.column_dimensions['B'].width = 35
    ws_patterns.column_dimensions['C'].width = 12
    ws_patterns.column_dimensions['D'].width = 80

    patterns_data = {
        "PT-001": ("Enterprise AI Governance", []),
        "PT-002": ("MLOps Level 2+", []),
        "PT-003": ("Feature Store", []),
        "PT-004": ("ML Explainability", []),
        "PT-005": ("Retrieval-Augmented Generation", []),
        "PT-006": ("Multi-Model Routing", []),
        "PT-007": ("Agentic AI", []),
        "PT-008": ("Multi-Agent Orchestration", []),
        "PT-009": ("Real-Time Scoring", []),
        "PT-010": ("Champion-Challenger", []),
        "PT-011": ("Intelligent Document Processing", []),
        "PT-012": ("Data Mesh", []),
        "PT-013": ("Conversational AI", []),
        "PT-014": ("Batch Prediction", []),
        "PT-015": ("Event-Driven Architecture", []),
        "PT-016": ("Stream Processing", []),
        "PT-017": ("Observability & Monitoring", []),
        "PT-018": ("Security & Privacy", []),
    }

    # Collect ABBs for each pattern
    for abb in ABBS:
        abb_id = abb[0]
        patterns = abb[6]  # Patterns are at index 6 (0=ID, 1=Name, 2=Category, 3=Desc, 4=Tech, 5=UC, 6=Patterns, 7=Layer)
        for pt in patterns.split(", "):
            pt = pt.strip()
            if pt in patterns_data:
                patterns_data[pt][1].append(abb_id)

    for row, (pt_id, (pt_name, abbs)) in enumerate(sorted(patterns_data.items()), 2):
        ws_patterns.cell(row=row, column=1, value=pt_id).border = thin_border
        ws_patterns.cell(row=row, column=2, value=pt_name).border = thin_border
        ws_patterns.cell(row=row, column=3, value=len(abbs)).border = thin_border
        ws_patterns.cell(row=row, column=4, value=", ".join(sorted(abbs))).border = thin_border
        ws_patterns.cell(row=row, column=4).alignment = cell_alignment

    # Save workbook
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    return len(ABBS)

def main():
    print("=" * 70)
    print("Generate ABB Catalog Excel Spreadsheet")
    print("=" * 70)
    print()

    count = create_excel()

    print(f"Created Excel file: {OUTPUT_FILE}")
    print(f"Total ABBs: {count}")
    print()
    print("Sheets created:")
    print("  1. ABB Catalog - Full catalog with all details including Architecture Layer")
    print("  2. Summary by Category - ABB counts by category")
    print("  3. Summary by Layer - ABB counts by architecture layer")
    print("  4. Pattern Coverage - ABBs mapped to each pattern")
    print()
    print("=" * 70)

if __name__ == "__main__":
    main()
